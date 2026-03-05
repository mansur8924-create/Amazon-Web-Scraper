import time  # Used to make the robot wait so it looks like a human
import random  # Used to pick random wait times and disguises so we aren't predictable
import logging  # The robot's diary; it writes down everything it does in the terminal
import pandas as pd  # The math engine that builds the Excel table and handles the data
import os  # The "construction crew" that creates folders on your computer
from pathlib import Path  # A smart way to handle file paths so Windows doesn't get confused
from datetime import datetime  # Used to put a "Date & Time" stamp on every price we find
from curl_cffi import requests  # The stealth browser that mimics real Google Chrome fingerprints
from bs4 import BeautifulSoup  # The "eyes" of the robot that read the HTML code of the website
import warnings  # Used to hide ugly, unimportant messages from appearing in your terminal

# Tells Python to ignore "FutureWarnings" so your terminal stays clean and professional
warnings.simplefilter(action='ignore')

# ---------------------------------------------------------
# 1. SETUP: Setting up the Robot's Diary
# ---------------------------------------------------------
logging.basicConfig(
    level=logging.INFO, # Sets the diary to record "Info" level events
    format='%(asctime)s - %(levelname)s - %(message)s', # Defines the look of the diary entry
    datefmt='%H:%M:%S' # Shows only the Hour:Minute:Second in the diary
)

# ---------------------------------------------------------
# 2. THE ENGINE: The Retail Intelligence Pipeline
# ---------------------------------------------------------
class RetailIntelligencePipeline:
    def __init__(self, target_excel: str):
        # Converts the string path into a smart Path object for easier handling
        self.target_excel = Path(target_excel)
        # The specific URL where Amazon ranks the most popular SSDs right now
        self.best_sellers_url = "https://www.amazon.ca/Best-Sellers-Electronics-Internal-Solid-State-Drives/zgbs/electronics/3312786011"
        # A list of fake "mustaches" (browser signatures) the robot can wear
        self.browser_profiles = ["chrome110", "chrome120", "edge101", "safari15_3", "safari15_5"]
        
        # A safety list of links to use if the Best Sellers page blocks us
        self.fallback_urls = [
            "https://www.amazon.ca/Samsung-990-PRO-SSD-2TB/dp/B0BHJJ9Y77/",
            "https://www.amazon.ca/WD_BLACK-SN850X-Internal-Gaming-Heatsink/dp/B0B7CKZHH9/",
            "https://www.amazon.ca/Crucial-Plus-PCIe-NAND-5000MB/dp/B0B25MJ1YT/",
            "https://www.amazon.ca/Sabrent-Internal-Extreme-Performance-SB-RKT4P-2TB/dp/B08P25PBJJ/"
        ]

    def _deploy_scout(self) -> list:
        """The Scout finds the current Top 10 links on the Best Sellers page."""
        logging.info("🦅 Deploying Scout to discover the current Top 10 Best Sellers...")
        # Randomly picks one browser disguise for the Scout to wear
        disguise = random.choice(self.browser_profiles)
        try:
            # Asks Amazon for the Best Sellers page using the disguise
            response = requests.get(self.best_sellers_url, impersonate=disguise, timeout=30)
            # Checks if Amazon showed us a "Robot Check" page instead of the products
            if "captcha" in response.text.lower() or "automated access" in response.text.lower():
                logging.warning("⚠️ Scout was blocked. Switching to Fallback targets.")
                return self.fallback_urls

            # Turns the raw website code into a "Soup" we can search through
            soup = BeautifulSoup(response.content, 'html.parser')
            # Finds every product box on the Best Sellers grid
            items = soup.find_all('div', {'id': 'gridItemRoot'})
            top_10_urls = []
            for item in items[:10]: # Loop through only the first 10 items found
                link_tag = item.find('a', class_='a-link-normal') # Find the clickable link
                if link_tag and 'href' in link_tag.attrs:
                    raw_url = "https://www.amazon.ca" + link_tag['href']
                    # Chops off tracking codes so the URL looks natural and clean
                    top_10_urls.append(raw_url.split('ref=')[0])
            # If the Scout found links, return them; otherwise, use the backup list
            return top_10_urls if top_10_urls else self.fallback_urls
        except Exception:
            return self.fallback_urls

    def _save_and_format_excel(self, new_data: list[dict]) -> None:
        """The Builder and Painter for the Excel file."""
        if not new_data: return # If no data was found, don't do anything
        # Converts our list of dictionary data into a clean Pandas Table
        df_new = pd.DataFrame(new_data)
        if self.target_excel.exists():
            # If the file already exists, read the old data
            df_existing = pd.read_excel(self.target_excel)
            # Glue the new data to the bottom of the old data
            df_combined = pd.concat([df_existing, df_new], ignore_index=True)
        else:
            # If it's a brand new file, the table only contains the new data
            df_combined = df_new
        # Save the whole table into the Excel file
        df_combined.to_excel(self.target_excel, index=False)
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            # Opens the Excel file we just saved to add the "Paint"
            wb = load_workbook(self.target_excel)
            ws = wb.active # Selects the active sheet
            header_fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid") # Dark Blue
            header_font = Font(color="FFFFFF", bold=True) # White and Bold
            # Loops through every cell in the first row to apply the header style
            for cell in ws[1]: 
                cell.fill, cell.font = header_fill, header_font
                cell.alignment = Alignment(horizontal="center")
            # Centers all the data in every row so it looks clean
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row: cell.alignment = Alignment(horizontal="center")
            # Measures the longest word in each column and widens the column to fit
            for col in ws.columns:
                max_len = max([len(str(cell.value)) for cell in col])
                ws.column_dimensions[col[0].column_letter].width = max_len + 5
            wb.save(self.target_excel) # Final save of the painted workbook
            print("🎨 SUCCESS: Dashboard formatted.")
        except Exception as e: logging.error(f"Excel error: {e}")

    def run_stealth_scan(self) -> None:
        """The Master process that runs the scout and then the scraper."""
        now = datetime.now().strftime('%Y-%m-%d %H:%M')
        print(f"\n[{now}] 🚀 Initiating Deep-Stealth Dynamic Scan...")
        target_urls = self._deploy_scout() # Runs the Scout first to find the targets
        gathered_intelligence = [] # An empty bag to hold the data we find

        for index, url in enumerate(target_urls, start=1):
            try:
                # MAX STEALTH: Waits between 45 and 90 seconds between each item
                sleep_timer = random.uniform(45, 90)
                logging.info(f"Deep Loitering active ({sleep_timer:.1f}s) for Target #{index}...")
                time.sleep(sleep_timer)

                # Goes to the product page with a fresh browser disguise
                response = requests.get(url, impersonate=random.choice(self.browser_profiles), timeout=30)
                if "captcha" in response.text.lower() or "automated access" in response.text.lower():
                    logging.error(f"🚨 Target #{index} blocked. Moving to next.")
                    continue

                soup = BeautifulSoup(response.content, 'html.parser')
                # Finds the product title
                title_elem = soup.find("span", {"id": "productTitle"})
                if not title_elem: continue # If the page didn't load properly, skip it
                
                name = title_elem.get_text().strip()
                # Finds the invisible price tag Amazon uses
                price = soup.find("span", {"class": "a-offscreen"})
                live_p = price.get_text().strip().replace('$', '').replace(',', '') if price else "N/A"
                
                # Grabs the star rating (e.g. 4.8)
                rating_elem = soup.find("span", {"class": "a-icon-alt"})
                rating = rating_elem.get_text().strip().split(' ')[0] if rating_elem else "N/A"
                # Grabs the total review count and cleans up the symbols
                review_elem = soup.find("span", {"id": "acrCustomerReviewText"})
                revs = review_elem.get_text().strip().split(' ')[0].replace(',', '').replace('(', '').replace(')', '') if review_elem else "0"

                # Adds all this info into a small "package"
                gathered_intelligence.append({
                    'Timestamp': now, 'Product': name[:60], 'Live Price': f"${live_p}",
                    'Rating': rating, 'Total Reviews': revs, 'Status': "In Stock"
                })
                print(f"✅ ACQUIRED: Target #{index} | ${live_p}")
            except Exception as e: logging.error(f"Failed on #{index}: {e}")
        
        # After scanning all 10, send the bag of data to the Excel Builder
        self._save_and_format_excel(gathered_intelligence)
        print("-" * 60 + "\n🏁 Scan Cycle Complete.")

# ---------------------------------------------------------
# 3. EXECUTION: The Heartbeat Loop
# ---------------------------------------------------------
if __name__ == "__main__":
    # The final destination address for your report
    DEST = r'C:\Users\mansu\OneDrive\Desktop\Data Analyst Boot Camp\Amazon web scraper(project)-python\SSD_Market_Master.xlsx'
    # Creates the folder if it's missing
    os.makedirs(os.path.dirname(DEST), exist_ok=True)
    pipeline = RetailIntelligencePipeline(target_excel=DEST)
    
    try:
        while True: # An infinite loop that never stops
            pipeline.run_stealth_scan() # Run the scan
            # Tell the robot to hibernate for exactly 3 hours (10,800 seconds)
            print(f"\n⏳ Entering 3-hour hibernation (Cycle finished at {datetime.now().strftime('%H:%M')})...")
            time.sleep(10800) 
    except KeyboardInterrupt:
        # If you press Ctrl + C, the robot wakes up and shuts down nicely
        print("\n🛑 Pipeline safely shut down.")
        
              