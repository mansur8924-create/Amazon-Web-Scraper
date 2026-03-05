"""
PROJECT: Retail Intelligence Pipeline (Amazon Web Scraper)
PURPOSE: Automated market monitoring for Internal Solid State Drives (SSDs).
FEATURES: 
    - Dynamic target discovery (Best Sellers list).
    - Stealth-based loitering to avoid IP blocking.
    - Automated Excel formatting with OpenPyXL.
    - 3-hour automated refresh cycle for time-series analysis.
AUTHOR: Mansur Mohammed
DATE: 2026-03-04
"""

import time
import random
import logging
import pandas as pd
from pathlib import Path
from datetime import datetime
from curl_cffi import requests
from bs4 import BeautifulSoup
import warnings

# Suppress minor data fragmentation warnings to maintain a clean console log
warnings.simplefilter(action='ignore')

# ---------------------------------------------------------
# 1. SETUP: System Logging (The Robot's Diary)
# ---------------------------------------------------------
# WHY: In a production environment, we need a record of successes and failures 
# without staring at the screen. This log provides a professional audit trail.
logging.basicConfig(
    level=logging.INFO, 
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%H:%M:%S'
)

# ---------------------------------------------------------
# 2. THE ENGINE: RetailIntelligencePipeline
# ---------------------------------------------------------
class RetailIntelligencePipeline:
    def __init__(self, target_excel: str):
        """
        Initializes the pipeline with target URLs and browser spoofing profiles.
        """
        # Using Pathlib ensures this script runs on Windows, Mac, or Linux seamlessly.
        self.target_excel = Path(target_excel)
        
        # Target: Amazon Canada - Best Sellers in Internal Solid State Drives
        self.best_sellers_url = "https://www.amazon.ca/Best-Sellers-Electronics-Internal-Solid-State-Drives/zgbs/electronics/3312786011"
        
        # 'The Disguises': Spoofing different browsers to prevent Amazon's anti-bot 
        # triggers from identifying the script as a scraper.
        self.browser_profiles = ["chrome110", "chrome120", "edge101", "safari15_3", "safari15_5"]
        
        # THE FALLBACK: If the dynamic 'Scout' is blocked, these high-volume 
        # targets ensure the data collection cycle doesn't return zero results.
        self.fallback_urls = [
            "https://www.amazon.ca/Samsung-990-PRO-SSD-2TB/dp/B0BHJJ9Y77/",
            "https://www.amazon.ca/WD_BLACK-SN850X-Internal-Gaming-Heatsink/dp/B0B7CKZHH9/",
            "https://www.amazon.ca/Crucial-Plus-PCIe-NAND-5000MB/dp/B0B25MJ1YT/",
            "https://www.amazon.ca/Sabrent-Internal-Extreme-Performance-SB-RKT4P-2TB/dp/B08P25PBJJ/"
        ]

    def _deploy_scout(self) -> list:
        """
        Step 1: The Scout. Dynamically discovers the current Top 10 products.
        WHY: Best-seller lists change daily; this ensures our data stays relevant.
        """
        logging.info("🦅 Deploying Scout to discover current Top 10 Market Leaders...")
        disguise = random.choice(self.browser_profiles)
        
        try:
            response = requests.get(self.best_sellers_url, impersonate=disguise, timeout=30)
            
            # Security Check: Detecting if we've been redirected to a Captcha page.
            if "captcha" in response.text.lower() or "automated access" in response.text.lower():
                logging.warning("⚠️ Scout detected Amazon security measures. Activating Fallback Protocol.")
                return self.fallback_urls

            soup = BeautifulSoup(response.content, 'html.parser')
            items = soup.find_all('div', {'id': 'gridItemRoot'})
            top_10_urls = []

            for item in items[:10]:
                link_tag = item.find('a', class_='a-link-normal')
                if link_tag and 'href' in link_tag.attrs:
                    raw_url = "https://www.amazon.ca" + link_tag['href']
                    # Clean the URL to remove tracking parameters for a cleaner dataset.
                    clean_url = raw_url.split('ref=')[0]
                    top_10_urls.append(clean_url)

            if top_10_urls:
                logging.info(f"📍 Scout successfully mapped {len(top_10_urls)} dynamic targets.")
                return top_10_urls
            else:
                logging.warning("⚠️ Scout encountered a layout change. Reverting to Fallback.")
                return self.fallback_urls
                
        except Exception as e:
            logging.error(f"Critical Scout Failure: {e}. Utilizing Fallback List.")
            return self.fallback_urls

    def _save_and_format_excel(self, new_data: list[dict]) -> None:
        """
        The Builder: Handles data persistence and applies professional formatting.
        WHY: Raw data is hard to read. A 'formatted' file is ready for executive review.
        """
        if not new_data:
            return 

        df_new = pd.DataFrame(new_data)
        
        # Data Persistence: If the file exists, append the new data to create a history.
        if self.target_excel.exists():
            df_existing = pd.read_excel(self.target_excel)
            df_combined = pd.concat([df_existing, df_new], ignore_index=True)
        else:
            df_combined = df_new

        df_combined.to_excel(self.target_excel, index=False)

        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Font, PatternFill, Alignment

            wb = load_workbook(self.target_excel)
            ws = wb.active

            # --- VISUAL STYLING (The Professional Touch) ---
            # Dark Blue Header with White Text (Industry Standard Branding)
            header_fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            for cell in ws[1]: 
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            # Column Auto-Adjustment: Ensures the Excel file is readable immediately upon opening.
            for col in ws.columns:
                max_length = 0
                column_letter = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[column_letter].width = max_length + 5

            wb.save(self.target_excel)
            logging.info("🎨 SUCCESS: Market Intelligence Dashboard formatted.")
        except PermissionError:
            logging.error("Formatting Failed: Please close the Excel file to allow the script to save.")
        except Exception as e:
            logging.error(f"Post-Processing Error: {e}")

    def run_stealth_scan(self) -> None:
        """
        The Main Mission: Navigates to each target and extracts granular data.
        """
        now = datetime.now().strftime('%Y-%m-%d %H:%M')
        print(f"\n[{now}] 🚀 Initiating Stealth Market Intelligence Scan...")
        print("-" * 80)

        target_urls = self._deploy_scout()
        gathered_intelligence = []

        for index, url in enumerate(target_urls, start=1):
            try:
                # 'Loitering': Randomized delays mimic human behavior and prevent IP flagging.
                sleep_timer = random.uniform(25, 45)
                logging.info(f"Stealth Active: Loitering for {sleep_timer:.1f}s before Target #{index}...")
                time.sleep(sleep_timer)

                disguise = random.choice(self.browser_profiles)
                response = requests.get(url, impersonate=disguise, timeout=30)
                
                if "captcha" in response.text.lower() or "automated access" in response.text.lower():
                    logging.error(f"🚨 HARD BLOCK: Amazon security triggered on Target #{index}.")
                    continue

                soup = BeautifulSoup(response.content, 'html.parser')
                title_elem = soup.find("span", {"id": "productTitle"})
                if not title_elem:
                    logging.warning(f"⚠️ Target #{index} returned a blank page. Skipping.")
                    continue
                
                # --- DATA EXTRACTION LOGIC ---
                product_name = title_elem.get_text().strip()
                
                # Extracting Price: Handling 'N/A' to prevent downstream math errors.
                price_element = soup.find("span", {"class": "a-offscreen"})
                live_price = price_element.get_text().strip().replace('$', '').replace(',', '') if price_element else "N/A"

                # Extracting Reviews and Ratings
                rating_elem = soup.find("span", {"class": "a-icon-alt"})
                rating = rating_elem.get_text().strip().split(' ')[0] if rating_elem else "N/A"

                review_elem = soup.find("span", {"id": "acrCustomerReviewText"})
                reviews = review_elem.get_text().strip().split(' ')[0].replace(',', '') if review_elem else "0"

                # Storing results in a dictionary for easy Pandas conversion
                gathered_intelligence.append({
                    'Timestamp': now,
                    'Product': product_name[:60] + "...", # Keep titles short for Excel readability
                    'Live Price': f"${live_price}" if live_price != "N/A" else "N/A",
                    'Rating': rating,
                    'Total Reviews': reviews
                })
                
                print(f"✅ DATA ACQUIRED: Target #{index} | Current Price: ${live_price}")

            except Exception as error:
                logging.error(f"System failure on Target #{index}: {error}")
        
        if gathered_intelligence:
            self._save_and_format_excel(gathered_intelligence)

        print("-" * 80)
        print("🏁 Stealth Scan Sequence Complete.")

# ---------------------------------------------------------
# 3. EXECUTION: Starting the Engine
# ---------------------------------------------------------
if __name__ == "__main__":
    # ENVIRONMENT-AGNOSTIC PATH:
    # Using a relative filename allows this to work on any machine that clones the repo.
    DESTINATION_FILE = 'SSD_Market_Master.xlsx'
    
    pipeline = RetailIntelligencePipeline(target_excel=DESTINATION_FILE)
    
    print("🤖 Stealth Retail Pipeline is ONLINE. Initializing automated 3-hour cycles...")
    
    try:
        while True:
            pipeline.run_stealth_scan()
            print(f"\n⏳ Cycle complete. Entering 3-hour stealth hibernation...")
            print("   (Ctrl + C to safely deactivate the engine)")
            time.sleep(10800) # 10,800 seconds = 3 Hours
            
    except KeyboardInterrupt:
        print("\n🛑 Manual Override: Pipeline successfully deactivated by user.")
